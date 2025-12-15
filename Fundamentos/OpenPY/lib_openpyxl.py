# Importa as classes Workbook e load_workbook da biblioteca openpyxl
from openpyxl import Workbook, load_workbook

# Carrega o arquivo Excel "Produtos.xlsx"
tabela = load_workbook("produtos.xlsx")

# Seleciona a aba ativa da planilha excel
aba_ativa = tabela.active

# Percorre todas as células da coluna C
for celula in aba_ativa["C"]:
    # Verifica se o valor da célula é "Serviço"
    if celula.value == "Serviço":
        # Obtém o número da linha onde o valor foi encontrado
        linha = celula.row
        
        # Atualiza o valor da coluna D na mesma "linha"
        aba_ativa[f"D{linha}"] = 1.5

# Salva as alterações em um novo arquivo Excel
tabela.save("produto_openpyxl.xlsx")


