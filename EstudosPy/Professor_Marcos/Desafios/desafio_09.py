# Importa as classes Workbook e load_workbook da biblioteca openpyxl
from openpyxl import Workbook, load_workbook

# Carrega o arquivo Excel "produtos.xlsx"
tabela = load_workbook("produtos.xlsx")

# Seleciona a aba ativa da planilha Excel
aba_ativa = tabela.active

# Percorre todas as células da coluna A
for celula in aba_ativa["A"]:
    # Verifica se o valor da célula da coluna A é "SPA"
    if celula.value == "SPA":
        # Caso encontre "SPA", percorre todas as células da coluna C
        for celula_c in aba_ativa["C"]:
            # Verifica se o valor da célula da coluna C é "Serviço"
            if celula_c.value == "Serviço":
                # Obtém o número da linha onde "Serviço" foi encontrado
                linha = celula_c.row
                
                # Atualiza o valor da coluna D na mesma linha
                aba_ativa[f"D{linha}"] = 2.1

# Salva as alterações em um novo arquivo Excel
tabela.save("produto_openpyxl.xlsx")
